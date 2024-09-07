import openpyxl as xl
from openpyxl.chart import BarChart, Reference

# Load the workbook and select the worksheet
wb = xl.load_workbook('Book1.xlsx')
sheet = wb['Sheet1']

# Loop through rows to apply correction and add values to the new column
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    if cell.value is not None:  # Check if cell.value is not None
        corrected_price = cell.value * 0.9
    else:
        corrected_price = 0  # Default value if cell.value is None
    
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

# Create a reference for the bar chart
values = Reference(sheet,
                   min_row=1,
                   max_row=sheet.max_row,
                   min_col=4,
                   max_col=4)

# Create and add the bar chart
chart = BarChart()
chart.add_data(values, titles_from_data=True)
sheet.add_chart(chart, 'E2')

# Save the modified workbook to a new file
wb.save('Book2.xlsx')
